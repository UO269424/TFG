import os
import numpy as np
import pandas as pd
from tensorflow import keras
from tensorflow.keras.preprocessing.image import load_img, img_to_array
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import Conv2D, MaxPooling2D, Flatten, Dense, TimeDistributed, LSTM, GRU
from tensorflow.keras.models import load_model
import openpyxl

# Ruta de la carpeta que contiene las imágenes
carpeta_imagenes = "C:/Users/Alonso/Desktop/Screenshots-Converted-v2"
carpeta_test_1 = "../ImageClassifier/Test"
carpeta_modelos = 'Modelos'
resultados_excel = 'Resultados/resultados.xlsx'
hiperparametros_csv = 'hiperparametros.csv'

# Parámetros de las imágenes
dim_imagen = (50, 50)
canales_color = 3


def cargar_imagenes_etiquetas(carpeta):
    nombres = sorted(os.listdir(carpeta))
    secuencias = []
    etiquetas = []

    for i in range(len(nombres) - 2):
        secuencia = []
        for j in range(3):
            nombre_archivo = nombres[i + j]
            ruta_imagen = os.path.join(carpeta, nombre_archivo)
            imagen = load_img(ruta_imagen, target_size=dim_imagen)
            imagen_array = img_to_array(imagen)
            imagen_array /= 255.0
            secuencia.append(imagen_array)
        secuencias.append(secuencia)

        if nombres[i + 2].endswith("_1.jpg"):
            etiqueta = 1  # "Copia"
        else:
            etiqueta = 0  # "No copia"
        etiquetas.append(etiqueta)

    x = np.array(secuencias)
    y = np.array(etiquetas)

    return x, y


def entrenar_modelo(x_train, y_train, hiperparametros):
    model = Sequential()
    model.add(TimeDistributed(Conv2D(int(hiperparametros['conv_filters_1']),
                                     (int(hiperparametros['conv_kernel_size']), int(hiperparametros['conv_kernel_size'])),
                                     strides=hiperparametros['strides'],
                                     padding='same',
                                     activation=hiperparametros['conv_activation']),
                              input_shape=(3, dim_imagen[0], dim_imagen[1], canales_color)))
    model.add(TimeDistributed(MaxPooling2D(pool_size=(int(hiperparametros['pool_size']), int(hiperparametros['pool_size'])),
                                           padding='valid')))
    if hiperparametros['conv_filters_2'] != '-':
        model.add(TimeDistributed(Conv2D(int(hiperparametros['conv_filters_2']),
                                         (int(hiperparametros['conv_kernel_size']), int(hiperparametros['conv_kernel_size'])),
                                         strides=hiperparametros['strides'],
                                         padding='same',
                                         activation=hiperparametros['conv_activation'])))
        model.add(TimeDistributed(MaxPooling2D(pool_size=(int(hiperparametros['pool_size']), int(hiperparametros['pool_size'])),
                                               padding='valid')))
    if hiperparametros['conv_filters_3'] != '-':
        model.add(TimeDistributed(Conv2D(int(hiperparametros['conv_filters_3']),
                                         (int(hiperparametros['conv_kernel_size']), int(hiperparametros['conv_kernel_size'])),
                                         strides=hiperparametros['strides'],
                                         padding='same',
                                         activation=hiperparametros['conv_activation'])))
        model.add(TimeDistributed(MaxPooling2D(pool_size=(int(hiperparametros['pool_size']), int(hiperparametros['pool_size'])),
                                               padding='valid')))
    model.add(TimeDistributed(Flatten()))
    if(hiperparametros['rnn'] == 'LSTM'):
        model.add(LSTM(int(hiperparametros['rnn_units']), activation=hiperparametros['rnn_activation']))
    if(hiperparametros['rnn'] == 'GRU'):
        model.add(GRU(int(hiperparametros['rnn_units']), activation=hiperparametros['rnn_activation']))
    model.add(Dense(1, activation='sigmoid'))

    model.compile(optimizer=hiperparametros['optimizer'], loss='binary_crossentropy', metrics=['accuracy'])

    model.fit(x_train, y_train, epochs=int(hiperparametros['epochs']))

    return model


def guardar_modelo(model, nombre_modelo):
    model.save(nombre_modelo)


def cargar_modelo(nombre_modelo):
    return load_model(nombre_modelo)


def evaluar_modelo(model, x_test, y_test):
    score = model.evaluate(x_test, y_test, verbose=1)
    return score


def escribir_resultados_excel(resultados, nombre_archivo):
    df = pd.DataFrame(resultados)
    if os.path.exists(nombre_archivo):
        book = openpyxl.load_workbook(nombre_archivo)
        writer = pd.ExcelWriter(nombre_archivo, mode='a', engine='openpyxl', if_sheet_exists= 'overlay')
        sheet = book.active
        startrow = sheet.max_row
        df.to_excel(writer, startrow=startrow, header=False, index=False)
        writer.close()
    else:
        df.to_excel(nombre_archivo, index=False)
"""

def escribir_resultados_excel(resultados, nombre_archivo):
    df = pd.DataFrame(resultados)
    if os.path.exists(nombre_archivo):
        wb = openpyxl.load_workbook(nombre_archivo)
        ws = wb.get_active_sheet()
        startrow = ws.max_row
        df.to_excel(wb, startrow=startrow, header=False, index=False)
"""
def leer_parametros_csv(nombre_archivo):
    parametros = pd.read_csv(nombre_archivo)
    return parametros


def obtener_ultimo_indice_modelo(nombre_carpeta_modelos):
    if not os.path.exists(nombre_carpeta_modelos):
        return 0
    modelos_existentes = os.listdir(nombre_carpeta_modelos)
    indices = [int(modelo.split('-')[1].split('.')[0]) for modelo in modelos_existentes if modelo.startswith('modelo-')]
    if(indices == []):
        return -1
    return max(indices)

def obtener_primer_indice_modelo(nombre_carpeta_modelos):
    if not os.path.exists(nombre_carpeta_modelos):
        return 0
    modelos_existentes = os.listdir(nombre_carpeta_modelos)
    indices = [int(modelo.split('-')[1].split('.')[0]) for modelo in modelos_existentes if modelo.startswith('modelo-')]
    if(indices == []):
        return -1
    return min(indices)


def obtener_precision_recall(model, x_test, y_test):
    # Realizar las predicciones del modelo en el conjunto de prueba
    y_pred = model.predict(x_test)
    y_pred = np.array(y_pred).flatten()
    y_pred = (y_pred > 0.5).astype(int)


    # Calcular True Positives, False Positives y False Negatives
    tp = np.sum(np.logical_and(y_pred == 1, y_test == 1))
    fp = np.sum(np.logical_and(y_pred == 1, y_test == 0))
    fn = np.sum(np.logical_and(y_pred == 0, y_test == 1))

    # Calcular precision y recall
    precision = tp / (tp + fp)
    recall = tp / (tp + fn)

    return precision, recall

def f_score(beta, precision, recall):
    f_beta = ((1 + beta**2)*((precision*recall)/((beta**2*precision)+recall)))
    return f_beta

def entrenamiento(carpeta_imagenes, hiperparametros_csv, carpeta_modelos):
    secuencias_train, etiquetas_train = cargar_imagenes_etiquetas(carpeta_imagenes)

    # Leer los hiperparámetros desde el archivo CSV
    parametros = leer_parametros_csv(hiperparametros_csv)

    ultimo_indice_modelo = obtener_ultimo_indice_modelo(carpeta_modelos)

    modelos_entrenados = []

    for i, hiperparametros in parametros.iterrows():
        modelo_indice = ultimo_indice_modelo + i + 1
        modelo_guardado = os.path.join(carpeta_modelos, 'modelo-' + str(modelo_indice) + '.h5')
        try:
            model = entrenar_modelo(secuencias_train, etiquetas_train, hiperparametros)
        except:
            print("Error en el modelo " + str(modelo_indice))
            continue;
        guardar_modelo(model, modelo_guardado)
        modelos_entrenados.append((modelo_guardado, modelo_indice))

    return modelos_entrenados

def probar_modelos(carpeta_modelos, carpeta_test_1, carpeta_test_2, carpeta_test_3, resultados_excel):
    parametros = leer_parametros_csv(hiperparametros_csv)
    resultados = []
    secuencias_test_1, etiquetas_test_1 = cargar_imagenes_etiquetas(carpeta_test_1)

    i = 0
    for nombre_archivo in os.listdir(carpeta_modelos):
        if nombre_archivo.endswith('.h5'):
            modelo_guardado = os.path.join(carpeta_modelos, nombre_archivo)
            modelo_indice = int(nombre_archivo.split('-')[1].split('.')[0])
            model = cargar_modelo(modelo_guardado)

            score_1 = evaluar_modelo(model, secuencias_test_1, etiquetas_test_1)
            precision_1 = obtener_precision_recall(model, secuencias_test_1, etiquetas_test_1)[0]
            recall_1 = obtener_precision_recall(model, secuencias_test_1, etiquetas_test_1)[1]
            metricas_1 = [score_1[1], precision_1, recall_1, f_score(1, precision_1, recall_1), f_score(2, precision_1, recall_1)]

            resultado = {
                'modelo': modelo_indice,
                'conv_filters_1': parametros.at[i,'conv_filters_1'],
                'conv_kernel_size': parametros.at[i,'conv_kernel_size'],
                'strides': parametros.at[i,'strides'],
                'pool_size': parametros.at[i,'pool_size'],
                'conv_filters_2': parametros.at[i,'conv_filters_2'],
                'conv_filters_3': parametros.at[i,'conv_filters_3'],
                'conv_activation': parametros.at[i, 'conv_activation'],
                'rnn': parametros.at[i,'rnn'],
                'rnn_units': parametros.at[i,'rnn_units'],
                'rnn_activation': parametros.at[i,'rnn_activation'],
                'optimizer': parametros.at[i,'optimizer'],
                'epochs': parametros.at[i,'epochs'],
                'accuracy-Test-1': metricas_1[0],
                'precision-Test-1': metricas_1[1],
                'recall-Test-1': metricas_1[2],
                'f1-Test-1': metricas_1[3],
                'f2-Test-1': metricas_1[4],
            }
            resultados.append(resultado)

            i+=1

            print('Modelo ' + str(modelo_indice) + ' metricas_1: ' + str(metricas_1))

    escribir_resultados_excel(resultados, resultados_excel)

def main():
    #modelos_entrenados = entrenamiento(carpeta_imagenes, hiperparametros_csv, carpeta_modelos)
    probar_modelos(carpeta_modelos, carpeta_test_1, carpeta_test_2, carpeta_test_3, resultados_excel)


if __name__ == '__main__':
    main()
